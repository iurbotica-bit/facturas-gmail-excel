import os
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS_MAPPING = {
    "fecha": "Fecha",
    "emisor": "Proveedor (Emisor)",
    "cif_emisor": "NIF/CIF Proveedor",
    "numero_factura": "Nº Factura",
    "base_imponible": "Base Imponible (€)",
    "iva_porcentaje": "IVA %",
    "iva_importe": "IVA Importe (€)",
    "total": "Total (€)",
    "archivo_origen": "Archivo Origen"
}

def get_trimestre_name(date_str: str) -> str:
    """Determina el nombre de la pestaña del trimestre basándose en la fecha (YYYY-MM-DD)."""
    if not date_str:
        # Si no hay fecha, lo asignamos al trimestre actual por defecto
        month = datetime.now().month
    else:
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            month = date_obj.month
        except Exception:
            month = datetime.now().month
            
    if month in [1, 2, 3]:
        return "1T"
    elif month in [4, 5, 6]:
        return "2T"
    elif month in [7, 8, 9]:
        return "3T"
    else:
        return "4T"

def apply_excel_styling(file_path: str):
    """Aplica un diseño estético premium a todas las hojas del Excel usando openpyxl."""
    wb = openpyxl.load_workbook(file_path)
    
    # Paleta de colores premium (Azul marino / Pizarra oscura)
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    
    border_side = Side(border_style="thin", color="E2E8F0")
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Estilo para la cabecera
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = cell_border
        
        # Altura de la fila de cabecera
        ws.row_dimensions[1].height = 28
        
        # Estilo para las filas de datos y formateo de números
        for r_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[r_idx].height = 20
            for c_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = data_font
                cell.border = cell_border
                
                # Alinear a la izquierda nombres, centro fechas e IDs, derecha números
                col_name = ws.cell(row=1, column=c_idx).value
                if col_name in ["Fecha", "NIF/CIF Proveedor", "Nº Factura"]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif " (€)" in str(col_name) or "%" in str(col_name):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    # Formato moneda o porcentaje
                    if cell.value is not None:
                        try:
                            cell.value = float(cell.value)
                            if "%" in str(col_name):
                                cell.number_format = '0.0"%"'
                            else:
                                cell.number_format = '#,##0.00" €"'
                        except ValueError:
                            pass
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
        # Auto-ajustar el ancho de las columnas
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.number_format and '€' in cell.number_format and isinstance(cell.value, (int, float)):
                    # Simular la longitud formateada
                    val_str = f"{cell.value:,.2f} €"
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    wb.save(file_path)

def save_to_excel(excel_path: str, data: dict, filename: str):
    """Guarda los datos extraídos en la hoja del trimestre correspondiente."""
    # Añadir archivo de origen a los datos
    data_to_save = data.copy()
    data_to_save["archivo_origen"] = filename
    
    # Determinar el trimestre
    trimestre = get_trimestre_name(data_to_save.get("fecha"))
    
    # Mapear las claves del diccionario a las columnas finales del Excel
    formatted_row = {}
    for raw_key, col_title in COLUMNS_MAPPING.items():
        formatted_row[col_title] = data_to_save.get(raw_key, None)
        
    df_new = pd.DataFrame([formatted_row])
    
    # Crear archivo si no existe
    if not os.path.exists(excel_path):
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_new.to_excel(writer, sheet_name=trimestre, index=False)
    else:
        # Cargar archivo existente
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Comprobar si la pestaña ya existe
            if trimestre in writer.sheets:
                # Leer datos existentes de esa pestaña
                df_existing = pd.read_excel(excel_path, sheet_name=trimestre)
                # Concatenar filas
                df_combined = pd.concat([df_existing, df_new], ignore_index=True)
                df_combined.to_excel(writer, sheet_name=trimestre, index=False)
            else:
                # Si no existe, crear la pestaña directamente
                df_new.to_excel(writer, sheet_name=trimestre, index=False)
                
    # Aplicar formato de tabla premium
    apply_excel_styling(excel_path)
